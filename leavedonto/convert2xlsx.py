from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import coordinate_to_tuple, get_column_letter


class Convert2Xlsx:
    def __init__(self, ont_path, ont):
        self.ont_path = ont_path
        self.ont = ont

    def convert2xlsx(self, out_path=None):
        def add_sheet(sheet, data, ft_style, starting_row=0):
            for row, entry in enumerate(data):
                for col, cell in enumerate(entry):
                    sheet.cell(starting_row + row + 1, col + 1).value = cell
                    sheet.cell(starting_row + row + 1, col + 1).font = ft_style
                    sheet.cell(starting_row + row + 1, col + 1).alignment = alignmnt

        tree = self.get_ont_tree()
        sheets = self.get_lists()
        font = 'Jomolhari'
        ft_structure = Font(font, size=15, color='0000CC')
        ft_legend = Font(font, size=15, color='0000CC')
        ft_entries = Font(font, size=15)
        alignmnt = Alignment(horizontal="left", vertical="top")

        wb = Workbook()
        wb.remove(wb['Sheet'])

        # adding the tree structure
        ws = wb.create_sheet('0 Ontology')

        add_sheet(ws, tree, ft_structure)
        self.resize_sheet(ws)
        ws.freeze_panes = 'B1'

        # adding the lists in individual sheets
        num = 1
        for title, sheet in sheets:
            ws = wb.create_sheet(f'{num} {title}')

            num += 1
            for n, lgd in enumerate(self.ont['legend']):
                ws.cell(1, n+1).value = lgd
                ws.cell(1, n + 1).font = ft_legend

            if not sheet:
                continue

            add_sheet(ws, sheet, ft_entries, starting_row=1)
            self.resize_sheet(ws)
            ws.freeze_panes = 'A2'

        if not out_path:
            out_path = self.ont_path.parent
        else:
            out_path = Path(out_path)

        if not out_path.is_dir():
            out_path.mkdir(exist_ok=True)

        out_file = Path(out_path) / (self.ont_path.stem + '.xlsx')

        wb.save(out_file)

    @staticmethod
    def resize_sheet(sheet):
        height = 20
        min_width = 5
        adjust = lambda x: 2 * x

        max_row, max_col = coordinate_to_tuple(sheet.dimensions.split(':')[1])
        # adjusting row heights
        for i in range(1, max_row+1):
            sheet.row_dimensions[i].height = height

        # adjusting col width
        cols = list(sheet.iter_cols())
        for i in range(max_col):
            col = cols[i]
            length = max([len(str(c.value)) for c in col])
            col_letter = get_column_letter(i + 1)
            sheet.column_dimensions[col_letter].width = adjust(length) if adjust(length) >= min_width else min_width

    def get_ont_tree(self):
        def extract_level(structure, level, key, value):
            if not value or isinstance(value, list):
                key += ':'
            structure.append([''] * level + [key])

        list_structure, level = [], 0
        self.__recursive_extract(self.ont['ont'], extract_level, list_structure, level)

        # add numbers
        idx = 1
        for n, l in enumerate(list_structure):
            to_add = ''
            if l[-1].endswith(':'):
                to_add = idx
                idx += 1
                list_structure[n] = l[:-1] + [l[-1][:-1]]  # remove the : at the end of the name

            list_structure[n] = [to_add] + l

        return list_structure

    def get_lists(self):
        def extract_lists(structure, _, key, value):
            if not isinstance(value, dict):
                structure.append((key, value))

        list_structure, level = [], 0
        self.__recursive_extract(self.ont['ont'], extract_lists, list_structure, level)
        return list_structure

    def __recursive_extract(self, ont, func, structure, level):
        for key, value in ont.items():
            func(structure, level, key, value)
            if isinstance(value, dict):
                self.__recursive_extract(value, func, structure, level + 1)
            else:
                continue
